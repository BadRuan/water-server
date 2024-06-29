from typing import List
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from core.settings import STATIONS, SF_COLOR, JJ_COLOR, BZ_COLOR
from core.model import Station, WaterLevel
from core.dao import (
    today8_waterlevel,
    yesterday8_waterlevel,
    lastweek8_waterlevel,
    lastyear8_waterlevel,
)


def filePath(source: str, dist: str):
    # 获取当前文件路径
    current_file_path = Path(__file__).resolve()
    # 获取当前文件的所在目录
    current_dir = current_file_path.parent.parent
    file_path = current_dir / f"source/{source}.xlsx"
    save_path = current_dir / f"dist/{dist}.xlsx"
    # 完整路径
    source_filename = str(file_path)
    dist_filename = str(save_path)
    return [source_filename, dist_filename]


# 获取目标水位
async def get_waterlevel_mode1() -> List[Station]:
    stations = STATIONS
    today8: List[WaterLevel] = await today8_waterlevel()
    yesterday_8: List[WaterLevel] = await yesterday8_waterlevel()
    lastweek8: List[WaterLevel] = await lastweek8_waterlevel()
    lastyear8: List[WaterLevel] = await lastyear8_waterlevel()

    def func(stations: List[Station], waterline_colum):
        for station in stations:  # 遍历每个测站
            for cow in waterline_colum:  # 每列数据
                if station.stcd == cow.stcd:
                    station.waterline.append(cow.current)

    for d in (today8, yesterday_8, lastweek8, lastyear8):
        func(stations, d)

    return stations


# 保存为xlsx文件
async def save_xlsx(
    source_file: str, save_file: str, talbe_loc: List[str], stations: List[Station]
):
    wb = load_workbook(source_file)
    ws = wb.active
    ws["A2"] = "填报日期： " + datetime.now().strftime("%Y年%m月%d日")

    def func(loc: str, index: int):
        for row, station in zip(ws[loc], stations):
            for cell in row:
                w_value = station.waterline[index]
                cell.value = w_value
                if w_value >= station.sfsw and w_value < station.jjsw:
                    cell.font = Font(color=SF_COLOR)
                elif w_value >= station.jjsw and w_value < station.bzsw:
                    cell.font = Font(color=JJ_COLOR)
                elif w_value >= station.bzsw:
                    cell.font = Font(color=BZ_COLOR)

    for index, loc in enumerate(talbe_loc):
        func(loc, index)

    wb.save(save_file)
