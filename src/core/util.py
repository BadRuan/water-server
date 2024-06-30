from typing import List
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from core.settings import SF_COLOR, JJ_COLOR, BZ_COLOR
from core.model import Station


def filePath(source: str, dist: str):
    # 获取当前文件路径
    current_file_path = Path(__file__).resolve()
    # 获取当前文件的所在目录
    current_dir = current_file_path.parent.parent
    file_path = current_dir / f"source/{source}.xlsx"
    save_path = current_dir / f"dist/{dist}.xlsx"
    # 完整路径
    source_filename = str(file_path)
    dist_filename = str(save_path)
    return [source_filename, dist_filename]


async def save_xlsx1(source_file: str, save_file: str, stations: List[Station]):
    wb = load_workbook(source_file)
    ws = wb.active
    ws["A2"] = "填报日期： " + datetime.now().strftime("%Y年%m月%d日")

    def func(index: int, talbe_loc: str):
        for row, station in zip(ws[talbe_loc], stations):
            for cell in row:
                w_value = station.waterline[index]
                cell.value = w_value
                if w_value >= station.sfsw and w_value < station.jjsw:
                    cell.font = Font(color=SF_COLOR)
                elif w_value >= station.jjsw and w_value < station.bzsw:
                    cell.font = Font(color=JJ_COLOR, bold=True)
                elif w_value >= station.bzsw:
                    cell.font = Font(color=BZ_COLOR, bold=True)

    talbe_loc = [
        "D5:D14",
        "E5:E14",
        "F5:F14",
        "G5:G14",
    ]
    for index, loc in enumerate(talbe_loc):
        func(index, loc)

    wb.save(save_file)


# 表格2 列头的标题 整点时间
def table2_str():
    def datefunc(date: datetime) -> str:
        date = date.date()
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        if date == today:
            return "今日"
        elif date == yesterday:
            return "昨日"
        else:
            return "***"

    now_hour = datetime.now()
    one_day_ago = now_hour - timedelta(days=1)

    now_str = datefunc(now_hour) + now_hour.strftime("%H时")
    one_day_ago_str = datefunc(one_day_ago) + one_day_ago.strftime("%H时")

    return (now_str, one_day_ago_str)


async def save_xlsx2(source_file: str, save_file: str, stations: List[Station]):
    wb = load_workbook(source_file)
    ws = wb.active
    ws["A2"] = "填报日期： " + datetime.now().strftime("%Y年%m月%d日%H时")
    date_lsit = table2_str()
    ws["D3"], ws["E3"] = date_lsit[0], date_lsit[1]

    def func(index: int, talbe_loc: str):
        for row, station in zip(ws[talbe_loc], stations):
            for cell in row:
                w_value = station.waterline[index]
                cell.value = w_value
                if w_value >= station.sfsw and w_value < station.jjsw:
                    cell.font = Font(color=SF_COLOR)
                elif w_value >= station.jjsw and w_value < station.bzsw:
                    cell.font = Font(color=JJ_COLOR, bold=True)
                elif w_value >= station.bzsw:
                    cell.font = Font(color=BZ_COLOR, bold=True)

    talbe_loc = [
        "D5:D14",
        "E5:E14",
    ]
    for index, loc in enumerate(talbe_loc):
        # print('index: ',index,'| loc: ', loc)
        func(index, loc)

    wb.save(save_file)


# 表格3 列头的标题 整点时间
def table3_str():
    def datefunc(date: datetime) -> str:
        date = date.date()
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)
        if date == today:
            return "今日"
        elif date == yesterday:
            return "昨日"
        else:
            return "***"

    now_hour = datetime.now()
    two_hour_ago = now_hour - timedelta(hours=2)
    four_hour_ago = now_hour - timedelta(hours=4)
    six_hour_ago = now_hour - timedelta(hours=6)
    eight_hour_ago = now_hour - timedelta(hours=8)
    ten_hour_ago = now_hour - timedelta(hours=10)

    now_str = datefunc(now_hour) + now_hour.strftime("%H时")
    two_hour_ago_str = datefunc(two_hour_ago) + two_hour_ago.strftime("%H时")
    four_hour_ago_str = datefunc(four_hour_ago) + four_hour_ago.strftime("%H时")
    six_hour_ago_str = datefunc(six_hour_ago) + six_hour_ago.strftime("%H时")
    eight_hour_ago_str = datefunc(eight_hour_ago) + eight_hour_ago.strftime("%H时")
    ten_hour_ago_str = datefunc(ten_hour_ago) + ten_hour_ago.strftime("%H时")
    return (
        now_str,
        two_hour_ago_str,
        four_hour_ago_str,
        six_hour_ago_str,
        eight_hour_ago_str,
        ten_hour_ago_str,
    )


async def save_xlsx3(source_file: str, save_file: str, stations: List[Station]):
    wb = load_workbook(source_file)
    ws = wb.active
    ws["A2"] = "填报日期： " + datetime.now().strftime("%Y年%m月%d日%H时")
    date_lsit = table3_str()
    ws["D3"] = date_lsit[0]
    ws["E3"] = date_lsit[1]
    ws["F3"] = date_lsit[2]
    ws["G3"] = date_lsit[3]
    ws["H3"] = date_lsit[4]
    ws["I3"] = date_lsit[5]

    def func(index: int, talbe_loc: str):
        for row, station in zip(ws[talbe_loc], stations):
            for cell in row:
                w_value = station.waterline[index]
                cell.value = w_value
                if w_value >= station.sfsw and w_value < station.jjsw:
                    cell.font = Font(color=SF_COLOR)
                elif w_value >= station.jjsw and w_value < station.bzsw:
                    cell.font = Font(color=JJ_COLOR, bold=True)
                elif w_value >= station.bzsw:
                    cell.font = Font(color=BZ_COLOR, bold=True)

    talbe_loc = [
        "D5:D14",
        "E5:E14",
        "F5:F14",
        "G5:G14",
        "H5:H14",
        "I5:I14",
    ]
    for index, loc in enumerate(talbe_loc):
        # print('index: ',index,'| loc: ', loc)
        func(index, loc)

    wb.save(save_file)
